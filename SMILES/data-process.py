"""
func: 生成数据
"""
import os
import csv
import openpyxl
from unit import Args, Log


def data_process(args: Args):
    log = Log(args.log)
    # 底物
    substrate = args.substrate
    # fasta 文件夹
    fasta_dir = args.fasta
    # 生成数据文件
    input_file = args.input_file

    # 读取底物数据
    workbook = openpyxl.load_workbook(substrate)
    sheet_names = workbook.sheetnames
    first_sheet = sheet_names[0]

    count = 1
    substrate_data = []
    # 拿到底物信息: [[name, smiles], ]
    for row in workbook[first_sheet].iter_rows(min_row=2, max_row=168, min_col=1, max_col=7):
        substrate_name = row[1].value
        substrate_smiles = row[3].value

        # 排除404
        if substrate_smiles == '404':
            substrate_smiles = row[6].value
        # 排除 None
        if substrate_smiles is None:
            continue

        # 共计113个
        log.append(f'{count}, {substrate_name}, {substrate_smiles}')
        count += 1

        substrate_data.append([substrate_name, substrate_smiles])\

    print(len(substrate_data))
    # raise

    # 拿到 fasta 序列, 共计198个
    # 获取fasta文件夹下所有文件名，并去除后缀，（过滤隐藏文件）
    fasta_files = [item[:-4] for item in os.listdir(fasta_dir) if not item.startswith('.')]
    print(len(fasta_files))

    # 生成数据文件
    with open(input_file, 'w', encoding='utf-8') as w:
        # w = csv.writer(w, delimiter='\t')
        # temp = "Substrate Name\tSubstrate SMILES\tProtein Sequence\tProtein EFI\n"
        temp = "Substrate Name#Substrate SMILES#Protein Sequence#Protein EFI\n"
        # temp = ['Substrate Name', "Substrate SMILES", "Protein Sequence", "Protein EFI"]
        w.write(temp)

        for efi_id in fasta_files:
            full_path = os.path.join(fasta_dir, f'{efi_id}.txt')
            with open(full_path, "r", encoding='utf-8') as r:
                fasta = r.read()
                fasta = fasta.strip()
                fasta = fasta.replace('\n', '')

            for substrate_item in substrate_data:
                temp = f'{substrate_item[0]}\t{substrate_item[1]}\t{fasta}\t{efi_id}\n'
                temp = f'{substrate_item[0]}#{substrate_item[1]}#{fasta}#{efi_id}\n'
                # temp = [substrate_item[0], substrate_item[1], fasta, efi_id]
                w.write(temp)

    pass


if __name__ == "__main__":
    data_process(Args())
    pass
