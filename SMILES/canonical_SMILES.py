"""
func: 从puhchem上爬底物的SMILES图

SMILES网址：https://pubchem.ncbi.nlm.nih.gov/compound/CDP#section=InChI-Key
XPath：//*[@id="Canonical-SMILES"]/div[2]/div[1]/p
改用selenium

chrome driver:
https://registry.npmmirror.com/binary.html?path=chromedriver/
"""
import os.path

import openpyxl

from selenium import webdriver
import selenium.common.exceptions as Exceptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from tqdm import tqdm

import requests


class Driver:
    def __init__(self):
        self.is_test_service = False

        self._wait_time = 6

        self._address = '/Users/yudd/Downloads/chromedriver'
        self._driver_is_create = True
        self.chrome = self._create()
        pass

    def _create(self):
        # 创建一个参数对象，用来控制chrome以无界面模式打开，反制selenium采取了监测机制，设置下载目录
        options = webdriver.ChromeOptions()
        # options.add_argument('--headless')
        options.add_argument('--ignore-ssl-errors=yes')
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--disable-gpu')
        options.add_experimental_option('excludeSwitches',
                                        ['enable-automation'])

        try:
            # 创建浏览器对象，Chrome驱动放在了Python一样的安装目录
            if self.is_test_service:
                driver = webdriver.Chrome(options=options)
            else:
                driver = webdriver.Chrome(executable_path=self._address, options=options)
            self._driver_is_create = True
        except Exceptions.SessionNotCreatedException as e:
            self._driver_is_create = False
            raise

        return driver

    def get(self, url):
        self.chrome.get(url)

    def find_element(self, find_type, find_value):
        try:
            return WebDriverWait(driver=self.chrome, timeout=self._wait_time, poll_frequency=0.5).until(
                EC.visibility_of_element_located((find_type, find_value))
            )
        except Exceptions.TimeoutException as _:
            return None

    def __del__(self):
        if self._driver_is_create:
            self.chrome.close()


class Args:
    def __init__(self):
        self._xlsx = 'pnas.1423570112.sd01.xlsx'
        if not os.path.exists(self._xlsx):
            print("Error: not found excel")
            raise

        # [name, url]
        self.substrate = self._data_preprocessing()

        # save result
        self.substrate_smiles = 'substrate_smiles.xlsx'

    def _data_preprocessing(self):
        substrate = []

        # read source excel file
        source_wordbook = openpyxl.load_workbook(self._xlsx, read_only=True)
        source_sheet_first_name = source_wordbook.sheetnames[0]
        source_sheet = source_wordbook[source_sheet_first_name]

        # for row in source_sheet.rows: first_col = row[2].value
        for row in source_sheet.iter_rows(min_row=2, max_row=168, min_col=3, max_col=3):
            data = []
            cell = str(row[0].value)
            cell = cell.strip()

            data.append(cell)

            char_index = cell.find('(')
            if char_index != -1:
                cell = cell[:char_index]

            url = f"https://pubchem.ncbi.nlm.nih.gov/compound/{cell}"

            data.append(url)

            substrate.append(data)
            # break

        return substrate


class Spider:
    def __init__(self):
        self._args = Args()
        self._substrate = self._args.substrate
        self._substrate_smiles_xlsx = self._args.substrate_smiles

        self._workbook = openpyxl.Workbook(write_only=True)
        self._sheet = self._workbook.create_sheet("sheet0")

        self._driver = Driver()

        self._headers = {
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="102", "Google Chrome";v="102"',
            'Referer': 'https://ydduong.blog.csdn.net/article/details/104409230',
            'sec-ch-ua-mobile': '?0',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.61 Safari/537.36',
            'sec-ch-ua-platform': '"macOS"',
            'Content-Type': 'text/plain;charset=UTF-8',
        }

        self._tqdm = None

    def _get_html(self, _url):
        """
        call url and return html
        :return: -1          -> find unknown error and will print url+error
                 443         -> find 'port=443' and need visit again
                 404         -> not find
                 _resp       -> return html object
        test:
        html = self._get_html(subs[1])

            if html == -1:
                subs.append("-1")
                continue

            if html == 404:
                subs.append("404")
                continue

            if html == 443:
                subs.append("443")
                continue

            print(html.text)
        """
        try:
            resp = requests.get(url=_url, headers=self._headers, timeout=(6, 6))
            if resp.status_code == 200:
                resp.encoding = resp.apparent_encoding
                return resp
            elif resp.status_code == 404:
                return 404
            else:
                print(f'Error: url: "{_url}"; status_code: {resp.status_code}')
                return -1
        except Exception as e:
            error_info = f'{e}'
            if error_info.find('port=443') != -1:
                return 443
            else:
                print(f'Error: url: "{_url}"; info: {e}')
                return -1

    def run(self):
        for subs in tqdm(self._substrate):
            url = subs[1]

            html = self._get_html(url)

            if html == 404:
                smiles = '404'
            else:
                # 访问网址
                self._driver.get(url)

                # 找到SMILES文本
                p = self._driver.find_element(By.XPATH, '//*[@id="Canonical-SMILES"]/div[2]/div[1]/p')
                if p is None:
                    smiles = 'None'
                else:
                    smiles = p.text

            subs.append(smiles)

            self._sheet.append(subs)

        self._workbook.save(self._substrate_smiles_xlsx)


if __name__ == '__main__':
    Spider().run()
