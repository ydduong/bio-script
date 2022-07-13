"""
SMILES网址：https://pubchem.ncbi.nlm.nih.gov/compound/CDP#section=InChI-Key
XPath：//*[@id="Canonical-SMILES"]/div[2]/div[1]/p
测试直接使用requests,无法获取元素
"""
import os.path

import requests
import time
import openpyxl


class Args:
    def __init__(self):
        self._xlsx = 'pnas.1423570112.sd01.xlsx'
        if not os.path.exists(self._xlsx):
            print("Error: not found excel")
            raise

        # [name, url]
        self.substrate = self._data_preprocessing()

    def _data_preprocessing(self):
        substrate = []

        # read source excel file
        source_wordbook = openpyxl.load_workbook(self._xlsx, read_only=True)
        source_sheet_first_name = source_wordbook.sheetnames[0]
        source_sheet = source_wordbook[source_sheet_first_name]

        # for row in source_sheet.rows: first_col = row[2].value
        for row in source_sheet.iter_rows(min_row=2, max_row=168, min_col=3, max_col=3):
            cell = str(row[0].value)
            cell = cell.strip()
            char_index = cell.find('(')
            if char_index != -1:
                cell = cell[:char_index]

            url = f"https://pubchem.ncbi.nlm.nih.gov/compound/{cell}"

            substrate.append([cell, url])
            # break

        return substrate


class Spider:
    def __init__(self):
        self._args = Args()
        self._substrate = self._args.substrate

        self._headers = {
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="102", "Google Chrome";v="102"',
            'Referer': 'https://ydduong.blog.csdn.net/article/details/104409230',
            'sec-ch-ua-mobile': '?0',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.61 Safari/537.36',
            'sec-ch-ua-platform': '"macOS"',
            'Content-Type': 'text/plain;charset=UTF-8',
        }

        self._tqdm = None

    def _get_html(self, _url):
        """
        call url and return html
        :return: -1          -> find unknown error and will print url+error
                 443         -> find 'port=443' and need visit again
                 404         -> not find
                 _resp       -> return html object
        test:
        html = self._get_html(subs[1])

            if html == -1:
                subs.append("-1")
                continue

            if html == 404:
                subs.append("404")
                continue

            if html == 443:
                subs.append("443")
                continue

            print(html.text)
        """
        try:
            resp = requests.get(url=_url, headers=self._headers, timeout=(6, 6))
            if resp.status_code == 200:
                resp.encoding = resp.apparent_encoding
                return resp
            elif resp.status_code == 404:
                return 404
            else:
                print(f'Error: url: "{_url}"; status_code: {resp.status_code}')
                return -1
        except Exception as e:
            error_info = f'{e}'
            if error_info.find('port=443') != -1:
                return 443
            else:
                print(f'Error: url: "{_url}"; info: {e}')
                return -1

    def run(self):
        for subs in self._substrate:
            html = self._get_html(subs[1])

            if html == -1:
                subs.append("-1")
                continue

            if html == 404:
                subs.append("404")
                continue

            if html == 443:
                subs.append("443")
                continue

            print(html.text)
            break

        time.sleep(1)


if __name__ == '__main__':
    Spider().run()
