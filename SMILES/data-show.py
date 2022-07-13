"""
func: 展示模型预测结果
"""
import os
import numpy
import openpyxl
from unit import Args
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image

# https://blog.csdn.net/z15818264727/article/details/122638802
from math import sqrt
from sklearn.metrics import mean_absolute_error
from sklearn.metrics import mean_squared_error
from sklearn.metrics import r2_score


def data_process(args: Args):
    # 处理模型输出数据，放入字典
    # {efi_id: {substrate_name: Kcat, }, }
    substrate_dict = dict()
    res_tsv = args.output_tsv
    with open(res_tsv, "r", encoding="utf-8") as r:
        lines = r.readlines()
        for line in lines:
            data = line.strip()
            data = data.split('\t')

            efi_id = data[-1]
            try:
                Kcat = float(data[-2])
            except:
                continue
            substrate_name = data[0]

            substrate_dict[efi_id] = substrate_dict.get(efi_id, dict())
            substrate_dict[efi_id][substrate_name] = Kcat

    # 删除表头
    # del substrate_dict['Kcat value (1/s)']
    # print(substrate_dict)

    # 数据序列
    in_order = ['blank', 'Phospho(enol)pyruvic acid (PEP)', '2-phosphoglyceric acid (glycerol 2-phosphate)', '3-phosphoglyceric acid (glycerol 3 phosphate)', 'D-erythronate-4-phosphate', 'L-erythronate-4-phosphate', 'D-threonate-4-phosphate', 'pyrophosphate (Ppi or POP)', 'ATP', 'TTP', 'GTP', 'dATP', 'Glycolic Acid-O-P', 'Glycerol-phosphate (GP)', 'DL-glyceraldehyde 3-phosphate', 'dihydroxyacetone phosphate (DHAP)', 'Meso-Erythritol-4-phosphate', 'D-Threitol-4-phosphate', 'erythrose-4-phosphate', 'imidodiphosphate (PNP)', 'ADP', 'TDP', 'GDP', 'dCTP', 'D-2-deoxy-Ribonate-5-phosphate', 'D-Arabinonate-5-phosphate', 'D-Lyxonate-5-phosphate', 'D-Ribonate-5-phosphate', 'D-Xylonate-3-phosphate', 'D-Xylonate-5-phosphate', 'L-Arabinonate-5-phosphate', 'thiamine pyrophosphate', 'UTP', 'CTP', 'ITP', 'dGTP', 'L-Ribonate-5-phosphate', 'L-Xylonate-3-phosphate', 'L-Xylonate-5-phosphate', 'D-Ribitol-5-phosphate', 'D-2-deoxy-ribitol-5-phosphate', 'D-Arabitol-1-P', 'D-Arabitol-4-P', "Uridine-5'-diphosphoglucuronic acid", 'UDP', 'CDP', 'IDP', 'dTTP', 'D-arabitol-5-phosphate', 'D-xylitol-3-phosphate', 'D-xylitol-5-phosphate', 'L-Arabitol-1-P', 'L-Arabitol-5-P', 'L-Lyxitol-5-P', 'L-Ribitol-5-P', 'sucrose 6-phosphate', "2'AMP", 'dCMP', 'GMP', 'Coenzyme A (CoA)', 'L-Xylitol-3-P', 'L-Xylitol-5-P', '2-deoxyribose 5-phosphate', 'arabinose 5-phosphate', 'D-Xylose-5-P', 'D-Ribose 5-phosphate', 'D-xylose-3-phosphate', 'trehalose-6-phosphate', "3'-AMP", 'dGMP', 'IMP', "pyridoxal-5'-phosphate", 'D-ribulose-5-phosphate', 'D-Lyxose-5-P', 'L-Arabinose-5-P', 'L-Lyxose-5-P', 'L-Ribose-5-P', 'L-Xylose-3-P', 'L-Xylose-5-P', '2-deoxy-6-Phosphogluconate', "5'-AMP", 'dUMP', 'TMP (dTMP)', 'riboflavin-5-phosphate (FMN)', 'L-ribulose-5-phosphate', 'D-xylulose-5-phosphate', 'L-Gulonic acid-2-Methylene Hydroxy phosphoric acid', 'D-sedoheptulose-7-phosphate', '2-deoxy-D-manno-2-octoulosonate-8-phosphate', '2-keto-3-deoxy-D-glycero-D-galactonononic acid-9-phosphate', 'N-acetylneuraminic acid-9-phosphate', '2-keto-3-deoxy-6-Phosphogluconate (KDPG)', 'dAMP', 'CMP', 'UMP', 'beta-Nicotinamide adenine dinucleotide phosphate (NADP)', 'D-3-deoxy-Gluconate-6-phosphate', 'D-Allonate-3-phosphate', 'D-Allonate-6-phosphate', 'D-Altronate-6-phosphate', 'D-Galactonate-6-phosphate', 'D-Gluconate-3-phosphate', 'D-Glucuronate-5-P', 'D-Glucuronic acid-5-P', "Adenosine 3',5'-diphosphate", 'DL-2-Amino-3-phosphonopropionic acid (APPA)', 'phosphonoformic acid (PFA)', 'N-phosphonomethyl glycine (PMG)', 'D-Mannonate-6-phosphate', 'L-Gluconate-3-P', '2-deoxy-6-phosphoglucitol', 'D-3-deoxy-sorbitol-6-phosphate', 'D-allitol-3-phosphate', 'D-allitol-6-phosphate', 'D-Galactitol-1-P', 'D-Galactitol-5-P', 'alpha-D-Glucose 1,6-bisphosphate', 'D-glycero-alpha-D-manno-heptose-1,7-bisphosphate', 'D-glucose-3,5-diphosphate', 'D-fructose 1,6-bisphosphate', 'D-galactitol-6-phosphate', 'D-Glucitol-2-P', 'D-glucitol-3-phosphate', 'D-Glucuronicitol-5-P', 'D-iditol-6-phosphate', 'D-Mannitol-1-P', 'D-Mannitol-2-P', 'D-Mannitol-4-P', 'beta-glucose-1,6-bisphosphate', 'D-glycero-beta-D-manno-heptose-1,7-bisphosphate', 'D-glucitol-3,5-diphohsphate', '5-phosphoribosyl-1-pyrophosphate', 'D-Mannitol-5-P', 'D-mannitol-6-phosphate', 'D-sorbitol-1-phosphate', 'L-Glucitol-3-P', '6-phosphogluconic acid', '2-Deoxy-D-glucose 6-phosphate', 'D-2-keto-glucose-6-phosphate', 'D-3-deoxy-glucose-6-phosphate', 'glucosamine 6 phosphate', 'O-phosphorylethanolamine', 'O-phospho-L-serine', 'acetyl-phosphate', 'D-allose-3-phosphate', 'D-allose-6-phosphate', 'D-Galactose-6-phosphate', 'D-glucose-3-phosphate', 'D-glucose-6-phosphate', 'D-mannose 6-phosphate', 'D-Mannose-2-P', 'L-Glucose-3-P', 'a-D-glucosamine-1-phosphate (not stable)', 'phosphocholine', 'O-phospho-L-threonine', 'Carbamyl phosphate', 'alpha D-galactose 1-phosphate', 'alpha-D-glucose-1-phosphate', 'alpha-D-mannose 1-phosphate', 'D-Fructose 6-phosphate', 'D-psicose-6-phosphate', 'D-tagatose-6-phosphate', 'L-sorbose-1-phosphate', 'L-Sorbose-4-P', 'N-acetyl-mannosamine-6-phosphate', 'Histidinol-phosphate (not stable)', 'O-phospho-L-tyrosine', 'p-nitrophenylphosphate (pNPP)']

    # 修改每张sheet表格
    r_2, rmse = [], []
    sheet_workbook = openpyxl.load_workbook(args.pnas_xlsx)
    res_sheet_name = sheet_workbook.sheetnames
    # name is efi id
    for sheet_name in res_sheet_name:
        if sheet_name not in substrate_dict.keys():
            continue
        curr_sheet = sheet_workbook[sheet_name]
        sheet_dict = substrate_dict[sheet_name]

        # 获取当前表格的酶活值（只获取有SMILES图的数据）
        activate_value = dict()
        order_index = 0
        for row in curr_sheet.iter_rows(min_row=7, max_row=20, min_col=3, max_col=14):
            for cell in row:
                if in_order[order_index] in sheet_dict.keys():
                    activate_value[in_order[order_index]] = cell.value
                order_index += 1

        # 将有SMILES图的结果数据附加到表格上来
        order_index = 0
        row_i, col_i = 24, 3
        while order_index < len(in_order):
            for num in range(12):
                curr_sheet.cell(row=row_i, column=col_i + num).value = sheet_dict.get(in_order[order_index], "")
                order_index += 1
            if order_index % 12 == 0:
                row_i += 1

        # 获取数据，做归一化
        label = []
        experiment_value = []
        predicted_value = []
        for key in sheet_dict.keys():
            label.append(key)
            experiment_value.append(activate_value[key])
            predicted_value.append(sheet_dict[key])

        # 都是108个
        # print(len(label), label)
        # print(len(experiment_value), experiment_value)
        # print(len(predicted_value), predicted_value)

        # 归一化
        max_exp = numpy.max(experiment_value)
        min_exp = numpy.min(experiment_value)
        range_exp = max_exp - min_exp
        for index, item in enumerate(experiment_value):
            experiment_value[index] = (item - min_exp) / range_exp

        max_exp = numpy.max(predicted_value)
        min_exp = numpy.min(predicted_value)
        range_exp = max_exp - min_exp
        for index, item in enumerate(predicted_value):
            predicted_value[index] = (item - min_exp) / range_exp

        # 计算评分
        temp_rmse = sqrt(mean_squared_error(experiment_value, predicted_value))
        temp_r_2 = r2_score(experiment_value, predicted_value)
        rmse.append(temp_rmse)
        r_2.append(temp_r_2)

        # 画图
        plt.figure()
        spot = plt.scatter(experiment_value, predicted_value)
        x = numpy.linspace(0, 1, 50)
        plt.plot(x, x, color='red', linewidth=2, linestyle='--')
        plt.xlim((0, 1))
        plt.ylim((0, 1))
        plt.xlabel('Predicted Kcat')
        plt.ylabel('Experiment value')
        plt.legend(handles=[spot, ], labels=[f'R2={temp_r_2}\nRMSE={temp_rmse}', ], loc="upper right")
        plt.title(f"EFI ID: {sheet_name}")
        pic = f'{sheet_name}.jpg'
        pic = os.path.join(args.efi_image, pic)
        plt.savefig(pic)
        plt.close()

        # 把图片保存到sheet中
        img = Image(pic)
        curr_sheet.add_image(img, "C40")

    # 画R2和RMSE图
    plt.figure()  # 设置画布的尺寸
    plt.title('R2 & RMSE boxplot')  # 标题，并设定字号大小
    plt.boxplot([r_2, rmse], labels=['R2', 'RMSE'], showmeans=True)
    pic = f'boxplot.jpg'
    pic = os.path.join(args.efi_image, pic)
    plt.savefig(pic)
    plt.close()

    second_sheet = sheet_workbook[res_sheet_name[1]]
    img = Image(pic)
    second_sheet.add_image(img, 'C21')

    sheet_workbook.save(args.pnas_kcat)


if __name__ == "__main__":
    data_process(Args())
    pass
