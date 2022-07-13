"""
func: Pans 是 读取表格数据
    : 外加爬虫获取对应的蛋白质序列
"""
import openpyxl
import os
import requests
from tqdm import tqdm

from unit import Args, Log


class Pnas:
    """
    func: 读取Excel文件信息: 获取Uniprot id信息
    """
    def __init__(self, args: Args):
        # 路径参数
        self._xlsx = args.pnas_xlsx

        # 加载文件
        self._workbook = openpyxl.load_workbook(filename=self._xlsx, read_only=True, data_only=True)
        # 获取页表
        self._sheet_names = self._workbook.sheetnames

        # 第一个页表记录，efi id 和 uniprot id 转换关系
        self._first_sheet = self._sheet_names[0]
        # 不同酶的efi id
        self._efi_id_sheet = self._sheet_names[3:]

        # key: efi 与 val: uniprot
        self._efi_uniprot_dict = dict()
        self._uniprot_efi_dict = dict()

        # key: uniprot 与 val: data
        self._uniprot_data_dict = dict()

        # init
        # 初始化 efi 与 uniprot 映射关系
        self._set_efi_uniprot_map()

    def _set_efi_uniprot_map(self):
        """
        func: 设置 efi 与 uniprot 映射关系
        """
        for row in self._workbook[self._first_sheet].iter_rows(min_row=2, min_col=1, max_col=2):
            if len(row) == 2:
                self._efi_uniprot_dict[str(row[0].value)] = str(row[1].value)
                self._uniprot_efi_dict[str(row[1].value)] = str(row[0].value)

    def get_efi_uniprot_map(self):
        """
        func: 获取 efi 与 uniprot 映射关系
        """
        return self._efi_uniprot_dict

    def get_uniprot_efi_map(self):
        """
        func: 获取 efi 与 uniprot 映射关系
        """
        return self._uniprot_efi_dict

    def get_uniprot_id(self):
        """
        func: 获取所有酶的Uniprot id
        note: 先获取所有有效页名（EFI ID），再根据对应关系，进行查找
        """
        _uniprot_id = []

        # 根据efi_id_sheet查找对应uniprot_id
        for item in self._efi_id_sheet:
            _uniprot_id.append(self._efi_uniprot_dict[item])

        return _uniprot_id

    def get_efi_id(self):
        return self._efi_id_sheet

    def get_uniprot_data_map(self):
        """
        func: 获取 uniprot id 和 data 之间的映射
        """
        for _efi_id in self._efi_id_sheet:
            # 每张表格数据
            _data = []
            for _row in self._workbook[_efi_id].iter_rows(min_row=7, max_row=20, min_col=3, max_col=14):
                for _cell in _row:
                    _data.append(_cell.value)

            _uniprot_id = self._efi_uniprot_dict[_efi_id]

            self._uniprot_data_dict[_uniprot_id] = _data

        return self._uniprot_data_dict

    def info(self):
        print(f'first sheet name: {self._first_sheet}')
        print(f'efi id names: {self._efi_id_sheet}')
        print(f'efi id number: {len(self._efi_id_sheet)}')

    def __del__(self):
        """
        info: 析构函数
        """
        # 关闭文件
        self._workbook.close()


def get_html(url):
    """
    func: call url and return html，html.text可以获取内容
    note: 如果状态码错误返回None, 如果是服务器访问过于频繁，拒绝访问，返回443，其他错误返回None
    """
    header = {
        'authority': 'cn.bing.com',
        'sec-ch-ua': '" Not;A Brand";v="99", "Google Chrome";v="97", "Chromium";v="97"',
        'sec-ch-ua-mobile': '?0',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36',
        'sec-ch-ua-arch': '"x86"',
        'sec-ch-ua-full-version': '"97.0.4692.71"',
        'sec-ch-ua-platform-version': '"10.0.0"',
        'sec-ch-ua-bitness': '"64"',
        'sec-ch-ua-model': '',
        'sec-ch-ua-platform': '"Windows"',
        'accept': 'image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-mode': 'no-cors',
        'sec-fetch-dest': 'image',
        'referer': 'https://cn.bing.com/?mkt=zh-CN',
        'accept-language': 'zh-CN,zh;q=0.9',
        'cookie': 'MUID=078F787B71F46FFD09CA68D070B76E50; MUIDB=078F787B71F46FFD09CA68D070B76E50; _EDGE_V=1; SRCHD=AF=NOFORM; SRCHUID=V=2&GUID=6ABD9AEC7BE248E2962C2E4E94E03E71&dmnchg=1; TTRSL=en; _tarLang=default=zh-Hans; _TTSS_OUT=hist=WyJlbiIsInpoLUhhbnMiXQ==; _TTSS_IN=hist=WyJpdCIsInJvIiwiZW4iLCJhdXRvLWRldGVjdCJd; BCP=AD=1&AL=1&SM=1; HOOKBLOCKINDICATOR=TRUE; ABDEF=V=13&ABDV=11&MRNB=1644823879884&MRB=0; _EDGE_S=SID=320E72F8C7BF611D2D9363B4C6B760CC&mkt=zh-cn; _SS=SID=320E72F8C7BF611D2D9363B4C6B760CC; SUID=M; ZHCHATSTRONGATTRACT=TRUE; ZHCHATWEAKATTRACT=TRUE; SRCHUSR=DOB=20210907&T=1644986902000&TPC=1644976209000; ipv6=hit=1644990503642&t=4; _UR=OMD=13289460522; SNRHOP=I=&TS=; _HPVN=CS=eyJQbiI6eyJDbiI6MTcsIlN0IjoyLCJRcyI6MCwiUHJvZCI6IlAifSwiU2MiOnsiQ24iOjE3LCJTdCI6MCwiUXMiOjAsIlByb2QiOiJIIn0sIlF6Ijp7IkNuIjoxNywiU3QiOjEsIlFzIjowLCJQcm9kIjoiVCJ9LCJBcCI6dHJ1ZSwiTXV0ZSI6dHJ1ZSwiTGFkIjoiMjAyMi0wMi0xNlQwMDowMDowMFoiLCJJb3RkIjowLCJHd2IiOjAsIkRmdCI6bnVsbCwiTXZzIjowLCJGbHQiOjAsIkltcCI6MjgyfQ==; SRCHHPGUSR=SRCHLANG=zh-Hans&BRW=NOTP&BRH=M&CW=332&CH=730&SW=1536&SH=864&DPR=1.25&UTC=480&DM=0&WTS=63780583702&HV=1644988141&BZA=0',
    }
    try:
        resp = requests.get(url=url, headers=header, timeout=(6, 6))
        resp.encoding = resp.apparent_encoding
        if resp.status_code == 200:
            return resp
        else:
            # print(f'url: {url} error code: {resp.status_code}')
            return None
    except Exception as e:
        error_info = f'{e}'
        if error_info.find('port=443') != -1:
            # print(f'{url} {e}')
            return 443
        else:
            # print(f'url: {url} error: {e}')
            return None


def spider(args: Args):
    """
    func: 根据Uniprot ID，爬取蛋白质序列，放在/data/fasta文件夹下
    note: 爬取前会先遍历fasta文件夹，若序列不存在，再爬取数据
    """
    # 路径参数
    fasta_dir = args.fasta

    # 日志文件
    log = Log(args.log)

    # 表格数据
    pnas = Pnas(args)
    # uniprot_efi_map
    uniprot_efi_dict = pnas.get_uniprot_efi_map()

    # 获取fasta文件夹下所有文件名，并去除后缀，（过滤隐藏文件）
    fasta_files = [item[:-4] for item in os.listdir(fasta_dir) if not item.startswith('.')]
    # 遍历Uniprot ID，并判断是不是已经存在，留下没有数据的
    uniprot_id = [item for item in pnas.get_efi_id() if item not in fasta_files]

    # 进度条
    bar = tqdm(total=len(uniprot_id), desc='spider processing')
    count = 0  # 计数，成功爬取的数目

    # 爬取信息
    for item in uniprot_id:
        bar.update(1)

        url = f'https://www.uniprot.org/uniprot/{item}.fasta'
        resp = get_html(url)
        if resp is not None:
            if resp == 443:
                # 服务器拒绝访问，返回443
                log.append(f'{item},{url},error: return 443')
            else:
                # 返回正常
                text = resp.text

                # 返回信息是否有效
                if len(text) == 0:
                    log.append(f'{item},{url},error: return html is Null')
                    continue

                # 去掉第一行提示信息
                text = text[text.find('\n') + 1:]

                # 保存信息
                efi_index = uniprot_efi_dict[item]
                file = os.path.join(fasta_dir, f'{efi_index}.txt')
                with open(file, "w", encoding="utf-8") as output:
                    output.write(text)
                    count += 1
        else:
            log.append(f'{item},{url},error: resp.status_code or exception')

    bar.close()
    info = f'共计{len(fasta_files)+len(uniprot_id)}条，已存在{len(fasta_files)}，待爬取{len(uniprot_id)}条，完成{count}条，失败{len(uniprot_id)-count}条'
    log.append(info)
    print(info)


if __name__ == '__main__':
    _args = Args()
    _pnas = Pnas(_args)
    _pnas.info()

    spider(_args)
