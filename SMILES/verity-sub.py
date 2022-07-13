"""
func: 验证分子底物顺序是否正确
    :
"""
import openpyxl
from unit import Args, Log
from openpyxl.drawing.image import Image


def verity(args: Args):
    substrate = args.substrate

    workbook = openpyxl.load_workbook(substrate)
    sheet_names = workbook.sheetnames
    first_sheet_name = sheet_names[0]
    second_sheet_name = sheet_names[1]

    substrate_name = dict()

    # 底物数量
    for row in workbook[first_sheet_name].iter_rows(min_row=2, max_row=168, min_col=2, max_col=2):
        content = row[0].value.strip()
        if content not in substrate_name.keys():
            substrate_name[content] = 0
        else:
            print("yy", content)
    print(len(substrate_name.keys()))

    # 自己对照的图
    in_order = []
    # [None, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
    for row in workbook[second_sheet_name].iter_rows(min_row=24, max_row=37, min_col=3, max_col=14):
        for cell in row:
            content = cell.value
            content = content.strip()

            if content not in substrate_name.keys():
                substrate_name[content] = 0
                print(f'kkkk: {content}')

            substrate_name[content] += 1

            in_order.append(content)

    # print(in_order)
    for item in substrate_name.keys():
        if substrate_name[item] != 1:
            print(item, substrate_name[item])

    # raise

    i, j = 40, 3
    print(substrate_name)
    substrate_name['L-Lyxitol-5-P'] = 88
    print(substrate_name)
    second_sheet = workbook[second_sheet_name]

    print(in_order)

    order_index = 0
    while order_index < len(in_order):
        for num in range(12):
            second_sheet.cell(row=i, column=j+num).value = in_order[order_index]
            order_index += 1
        if order_index % 12 == 0:
            i += 1

    # 验证两个字典的key是否一致
    print(len(substrate_name.keys()), len(in_order))
    for item in substrate_name.keys():
        if item not in in_order:
            print(item)

    temp = ['L-Xylitol-5-P', 'D-Galactitol-1-P', 'D-Mannitol-1-P', '2-deoxyribose 5-phosphate', 'D-Xylose-5-P', 'L-Xylose-5-P', 'D-Lyxose-5-P', 'UDP', "Uridine-5'-diphosphoglucuronic acid", 'UMP', 'GMP', 'alpha-D-Glucose 1,6-bisphosphate']
    for item in temp:
        if item not in in_order:
            print("kk", item)

    # 测试追加图片
    pic = "./temp.jpg"
    img = Image(pic)
    second_sheet.add_image(img, 'C58')

    workbook.save(substrate)


if __name__ == "__main__":
    verity(Args())
